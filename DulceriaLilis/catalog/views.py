from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from .models import Product, Category, Brand
from .forms import ProductForm, CategoryForm, BrandForm
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from django.utils.decorators import method_decorator
from users.decorators import group_required

# Product Views
@method_decorator(group_required('Administrador', 'Editor', 'Lector'), name='dispatch')
class ProductListView(ListView):
    model = Product
    template_name = 'catalog/product_list.html'
    context_object_name = 'products'

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(name__icontains=query) |
                Q(description__icontains=query)
            )
        
        sort_by = self.request.GET.get('sort_by', 'name')
        direction = self.request.GET.get('direction', 'asc')
        if direction == 'desc':
            sort_by = f'-{sort_by}'
        
        return queryset.order_by(sort_by)

    def get_paginate_by(self, queryset):
        paginate_by = self.request.GET.get('paginate_by', self.request.session.get('paginate_by', 10))
        self.request.session['paginate_by'] = paginate_by
        return paginate_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['paginate_by'] = int(self.get_paginate_by(self.get_queryset()))
        context['sort_by'] = self.request.GET.get('sort_by', 'name')
        context['direction'] = self.request.GET.get('direction', 'asc')
        return context

@method_decorator(group_required('Administrador', 'Editor', 'Lector'), name='dispatch')
class ProductDetailView(DetailView):
    model = Product
    template_name = 'catalog/product_detail.html'
    context_object_name = 'product'

@method_decorator(group_required('Administrador', 'Editor'), name='dispatch')
class ProductCreateView(CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'catalog/product_form.html'
    success_url = reverse_lazy('product_list')

@method_decorator(group_required('Administrador', 'Editor'), name='dispatch')
class ProductUpdateView(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'catalog/product_form.html'
    success_url = reverse_lazy('product_list')

@method_decorator(group_required('Administrador'), name='dispatch')
class ProductDeleteView(DeleteView):
    model = Product
    template_name = 'catalog/product_confirm_delete.html'
    success_url = reverse_lazy('product_list')

@group_required('Administrador', 'Editor', 'Lector')
def export_products_to_excel(request):
    products = Product.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Add headers
    headers = ["SKU", "Name", "Description", "Price"]
    ws.append(headers)

    # Add data
    for product in products:
        ws.append([product.sku, product.name, product.description, product.price])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)

    return response

# Category Views
@method_decorator(group_required('Administrador', 'Editor', 'Lector'), name='dispatch')
class CategoryListView(ListView):
    model = Category
    template_name = 'catalog/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(name__icontains=query)
        
        sort_by = self.request.GET.get('sort_by', 'name')
        direction = self.request.GET.get('direction', 'asc')
        if direction == 'desc':
            sort_by = f'-{sort_by}'
        
        return queryset.order_by(sort_by)

    def get_paginate_by(self, queryset):
        paginate_by = self.request.GET.get('paginate_by', self.request.session.get('paginate_by', 10))
        self.request.session['paginate_by'] = paginate_by
        return paginate_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['paginate_by'] = int(self.get_paginate_by(self.get_queryset()))
        context['sort_by'] = self.request.GET.get('sort_by', 'name')
        context['direction'] = self.request.GET.get('direction', 'asc')
        return context

@method_decorator(group_required('Administrador', 'Editor', 'Lector'), name='dispatch')
class CategoryDetailView(DetailView):
    model = Category
    template_name = 'catalog/category_detail.html'
    context_object_name = 'category'

@method_decorator(group_required('Administrador', 'Editor'), name='dispatch')
class CategoryCreateView(CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'catalog/category_form.html'
    success_url = reverse_lazy('category_list')

@method_decorator(group_required('Administrador', 'Editor'), name='dispatch')
class CategoryUpdateView(UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'catalog/category_form.html'
    success_url = reverse_lazy('category_list')

@method_decorator(group_required('Administrador'), name='dispatch')
class CategoryDeleteView(DeleteView):
    model = Category
    template_name = 'catalog/category_confirm_delete.html'
    success_url = reverse_lazy('category_list')

@group_required('Administrador', 'Editor', 'Lector')
def export_categories_to_excel(request):
    categories = Category.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categories"

    headers = ["Name", "Created At", "Updated At"]
    ws.append(headers)

    for category in categories:
        ws.append([
            category.name,
            category.created_at.strftime("%Y-%m-%d %H:%M:%S") if category.created_at else '',
            category.updated_at.strftime("%Y-%m-%d %H:%M:%S") if category.updated_at else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categories.xlsx'
    wb.save(response)

    return response


# Brand Views
@method_decorator(group_required('Administrador', 'Editor', 'Lector'), name='dispatch')
class BrandListView(ListView):
    model = Brand
    template_name = 'catalog/brand_list.html'
    context_object_name = 'brands'

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(name__icontains=query)
        
        sort_by = self.request.GET.get('sort_by', 'name')
        direction = self.request.GET.get('direction', 'asc')
        if direction == 'desc':
            sort_by = f'-{sort_by}'
        
        return queryset.order_by(sort_by)

    def get_paginate_by(self, queryset):
        paginate_by = self.request.GET.get('paginate_by', self.request.session.get('paginate_by', 10))
        self.request.session['paginate_by'] = paginate_by
        return paginate_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['paginate_by'] = int(self.get_paginate_by(self.get_queryset()))
        context['sort_by'] = self.request.GET.get('sort_by', 'name')
        context['direction'] = self.request.GET.get('direction', 'asc')
        return context

@method_decorator(group_required('Administrador', 'Editor', 'Lector'), name='dispatch')
class BrandDetailView(DetailView):
    model = Brand
    template_name = 'catalog/brand_detail.html'
    context_object_name = 'brand'

@method_decorator(group_required('Administrador', 'Editor'), name='dispatch')
class BrandCreateView(CreateView):
    model = Brand
    form_class = BrandForm
    template_name = 'catalog/brand_form.html'
    success_url = reverse_lazy('brand_list')

@method_decorator(group_required('Administrador', 'Editor'), name='dispatch')
class BrandUpdateView(UpdateView):
    model = Brand
    form_class = BrandForm
    template_name = 'catalog/brand_form.html'
    success_url = reverse_lazy('brand_list')

@method_decorator(group_required('Administrador'), name='dispatch')
class BrandDeleteView(DeleteView):
    model = Brand
    template_name = 'catalog/brand_confirm_delete.html'
    success_url = reverse_lazy('brand_list')

@group_required('Administrador', 'Editor', 'Lector')
def export_brands_to_excel(request):
    brands = Brand.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Brands"

    headers = ["Name", "Created At", "Updated At"]
    ws.append(headers)

    for brand in brands:
        ws.append([
            brand.name,
            brand.created_at.strftime("%Y-%m-%d %H:%M:%S") if brand.created_at else '',
            brand.updated_at.strftime("%Y-%m-%d %H:%M:%S") if brand.updated_at else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=brands.xlsx'
    wb.save(response)

    return response